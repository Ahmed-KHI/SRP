"""
Excel Exporter - Automated Excel Integration

This module provides functionality to export processed receipts and expense data
to Excel spreadsheets with professional formatting and charts.
"""

import logging
from typing import List, Optional, Dict, Any
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from models.receipt import ProcessedReceipt
from models.expense import ExpenseReport
from utils.config import Config

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Professional Excel export functionality for expense data."""
    
    def __init__(self, config: Config):
        """Initialize the Excel exporter."""
        self.config = config
        self.template_path = getattr(config, 'DEFAULT_EXCEL_TEMPLATE', None)
        
        # Styling configuration
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.data_font = Font(name='Arial', size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        logger.info("Excel exporter initialized")
    
    def export_receipts(self, receipts: List[ProcessedReceipt], output_path: str) -> bool:
        """
        Export processed receipts to Excel file.
        
        Args:
            receipts: List of processed receipts
            output_path: Path for the output Excel file
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Create main data sheet
            self._create_data_sheet(wb, receipts)
            
            # Create summary sheet
            self._create_summary_sheet(wb, receipts)
            
            # Create charts sheet
            self._create_charts_sheet(wb, receipts)
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Successfully exported {len(receipts)} receipts to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export receipts to Excel: {str(e)}")
            return False
    
    def export_to_template(self, receipts: List[ProcessedReceipt], template_path: str, output_path: str) -> bool:
        """
        Export receipts using an existing Excel template.
        
        Args:
            receipts: List of processed receipts
            template_path: Path to Excel template file
            output_path: Path for output file
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            # Load template
            wb = load_workbook(template_path)
            
            # Find data sheet (assume first sheet or 'Data' sheet)
            if 'Data' in wb.sheetnames:
                ws = wb['Data']
            else:
                ws = wb.active
            
            # Clear existing data (keep headers)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            
            # Add receipt data
            for row_idx, receipt in enumerate(receipts, start=2):
                data = receipt.to_excel_row()
                col_idx = 1
                
                for header_cell in ws[1]:
                    header = header_cell.value
                    if header in data:
                        ws.cell(row=row_idx, column=col_idx, value=data[header])
                    col_idx += 1
            
            # Save updated template
            wb.save(output_path)
            logger.info(f"Exported {len(receipts)} receipts using template to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export using template: {str(e)}")
            return False
    
    def export_dataframe(self, df: pd.DataFrame, output_path: str, sheet_name: str = "Data") -> bool:
        """
        Export pandas DataFrame to Excel with formatting.
        
        Args:
            df: DataFrame to export
            output_path: Output Excel file path
            sheet_name: Name of the Excel sheet
            
        Returns:
            True if successful, False otherwise
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format the sheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                self._format_worksheet(worksheet, len(df.columns))
            
            logger.info(f"Exported DataFrame to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export DataFrame: {str(e)}")
            return False
    
    def _create_data_sheet(self, workbook: Workbook, receipts: List[ProcessedReceipt]):
        """Create the main data sheet with all receipt details."""
        ws = workbook.create_sheet("Receipt Data")
        
        # Headers
        headers = [
            'Date', 'Vendor', 'Amount', 'Category', 'Description',
            'Account Code', 'Department', 'Status', 'Requires Review',
            'Confidence Score', 'Notes', 'Processing Date', 'Image File'
        ]
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Add data
        for row_idx, receipt in enumerate(receipts, 2):
            data = receipt.to_excel_row()
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=data.get(header, ''))
                cell.font = self.data_font
                cell.border = self.border
                
                # Format amount column
                if header == 'Amount' and isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0.00'
                
                # Format percentage
                if header == 'Confidence Score':
                    cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, workbook: Workbook, receipts: List[ProcessedReceipt]):
        """Create summary sheet with aggregated data."""
        ws = workbook.create_sheet("Summary")
        
        # Calculate summary statistics
        total_amount = sum(r.amount or 0 for r in receipts)
        total_count = len(receipts)
        avg_amount = total_amount / total_count if total_count > 0 else 0
        
        # Category breakdown
        categories = {}
        vendors = {}
        status_counts = {}
        
        for receipt in receipts:
            # Categories
            cat = receipt.category
            categories[cat] = categories.get(cat, 0) + (receipt.amount or 0)
            
            # Vendors
            vendor = receipt.vendor or "Unknown"
            vendors[vendor] = vendors.get(vendor, 0) + (receipt.amount or 0)
            
            # Status
            status = receipt.status
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Add summary data
        row = 1
        
        # Overall summary
        self._add_summary_section(ws, "EXPENSE SUMMARY", [
            ("Total Expenses", f"${total_amount:,.2f}"),
            ("Number of Receipts", total_count),
            ("Average Amount", f"${avg_amount:.2f}"),
            ("Requires Review", sum(1 for r in receipts if r.requires_review))
        ], row)
        
        row += 8
        
        # Category breakdown
        category_data = [(cat, f"${amt:,.2f}") for cat, amt in sorted(categories.items(), key=lambda x: x[1], reverse=True)]
        self._add_summary_section(ws, "BY CATEGORY", category_data, row)
        
        row += len(category_data) + 5
        
        # Top vendors
        top_vendors = sorted(vendors.items(), key=lambda x: x[1], reverse=True)[:10]
        vendor_data = [(vendor, f"${amt:,.2f}") for vendor, amt in top_vendors]
        self._add_summary_section(ws, "TOP VENDORS", vendor_data, row)
    
    def _create_charts_sheet(self, workbook: Workbook, receipts: List[ProcessedReceipt]):
        """Create charts sheet with visual analytics."""
        ws = workbook.create_sheet("Charts")
        
        # Prepare data for charts
        categories = {}
        for receipt in receipts:
            cat = receipt.category
            categories[cat] = categories.get(cat, 0) + (receipt.amount or 0)
        
        # Create category data table for charts
        row = 2
        ws.cell(row=row, column=1, value="Category").font = self.header_font
        ws.cell(row=row, column=2, value="Amount").font = self.header_font
        
        for i, (category, amount) in enumerate(categories.items(), 3):
            ws.cell(row=i, column=1, value=category)
            ws.cell(row=i, column=2, value=amount)
        
        # Create pie chart
        pie_chart = PieChart()
        pie_chart.title = "Expenses by Category"
        
        # Data for pie chart
        data_range = Reference(ws, min_col=2, min_row=2, max_row=len(categories) + 2)
        category_range = Reference(ws, min_col=1, min_row=3, max_row=len(categories) + 2)
        
        pie_chart.add_data(data_range, titles_from_data=True)
        pie_chart.set_categories(category_range)
        
        # Add chart to sheet
        ws.add_chart(pie_chart, "D2")
        
        # Create bar chart
        bar_chart = BarChart()
        bar_chart.title = "Category Spending"
        bar_chart.x_axis.title = "Categories"
        bar_chart.y_axis.title = "Amount ($)"
        
        bar_chart.add_data(data_range, titles_from_data=True)
        bar_chart.set_categories(category_range)
        
        ws.add_chart(bar_chart, "D18")
    
    def _add_summary_section(self, worksheet, title: str, data: List[tuple], start_row: int):
        """Add a summary section to the worksheet."""
        # Add title
        title_cell = worksheet.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Add data
        for i, (label, value) in enumerate(data, start_row + 2):
            worksheet.cell(row=i, column=1, value=label).font = Font(name='Arial', size=10, bold=True)
            worksheet.cell(row=i, column=2, value=value).font = Font(name='Arial', size=10)
    
    def _format_worksheet(self, worksheet, num_columns: int):
        """Apply standard formatting to a worksheet."""
        # Format header row
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
